#!/usr/bin/env python3
from __future__ import annotations
import base64, hashlib, io, json, os, re, secrets, sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

import requests
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from google.oauth2 import service_account
from google.auth.transport.requests import AuthorizedSession
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
PUBLIC_OUT=ROOT/'data'/'work-system.json'
ENCRYPTED_OUT=ROOT/'data'/'work-system.enc.json'
PBKDF2_ITERATIONS=600_000
PASSPHRASE=os.getenv('DASHBOARD_PASSPHRASE','')
SERVICE_ACCOUNT_JSON=os.getenv('GOOGLE_SERVICE_ACCOUNT_JSON','').strip()
OPENAI_API_KEY=os.getenv('OPENAI_API_KEY','').strip()
OPENAI_MODEL=os.getenv('OPENAI_MODEL','').strip()

SOURCES=[
    {'id':'19gBiQuP8kjMRkzhuNRXXKKZS7Cn-yKJUZdADPOYJDos','kind':'experience','label':'工作经验 / 工作体系'},
    {'id':'13nPr94W9YHfZi_SkFk1i38WAHdO8pdEq','kind':'book','label':'读书笔记'},
    {'id':'1bBVlTpIodFbSLK8l2CsukALFdSq9m5hpxLHMK7liaug','kind':'field_manual','label':'EC / Amazon 实战手册'},
]

DOMAINS=[
 ('decision','判断与问题定义',r'目的|違和感|判断|根拠|先入観|比較|情報源|課題|仮説|意思決定'),
 ('career','职业 / 转职 / Career',r'転職|转职|キャリア|職務経歴|職歴|面接|面談|オファー|求人|年収|給与|待遇|退職|入社|志望動機|職種|雇用|採用'),
 ('market','市场 / GTM / Positioning',r'市場|GTM|ポジショニング|代替|競合|差別|サブカテゴリー|トレンド|参入'),
 ('consumer','消费者 / 品牌 / CEP',r'消費者|生活者|顧客|購買|想起|CEP|ブランド|浸透|ロイヤ|態度|パーセプション'),
 ('product','商品 / 新品上市',r'新商品|新製品|商品開発|発売|ローンチ|SKU|価格|値上げ|値下げ'),
 ('ec','EC / Amazon / 流通',r'Amazon|楽天|EC|モール|検索|レビュー|セール|クーポン|在庫|出荷|CVR|コンバージョン'),
 ('media','广告 / PR / 内容',r'広告|PR|メディア|SNS|TikTok|UGC|プロモーション|キャンペーン|ROAS|ROI|リーチ'),
 ('data','数据 / KPI / 测量',r'データ|分析|KPI|KGI|指標|測定|数字|予測|記録|パラメータ|ROI'),
 ('management','执行 / 管理 / Stakeholder',r'マネジメント|上司|メンバー|チーム|依頼|共有|レビュー|交渉|ベンダー|パートナー|採用|関係者|実行'),
 ('ai','AI / 工作效率',r'AI|生成AI|LLM|自動|効率|録音|議事録'),
 ('customer','客户服务 / 体验',r'カスタマー|問い合わせ|返品|交換|評価|体験|CX|離脱'),
 ('overseas','海外 / 渠道扩张',r'海外|中国|米国|グローバル|越境'),
]

def clean(v):
    if v is None:return ''
    s=str(v).strip()
    return re.sub(r'\s+',' ',s)

def domain_tags(text):
    out=[]
    for key,name,pat in DOMAINS:
        if re.search(pat,text or '',re.I):out.append({'id':key,'name':name})
    return out[:5] or [{'id':'other','name':'其他'}]

def credentials_session():
    if not SERVICE_ACCOUNT_JSON:
        raise RuntimeError('GOOGLE_SERVICE_ACCOUNT_JSON is not configured.')
    try: info=json.loads(SERVICE_ACCOUNT_JSON)
    except json.JSONDecodeError as e: raise RuntimeError(f'GOOGLE_SERVICE_ACCOUNT_JSON is invalid JSON: {e}')
    creds=service_account.Credentials.from_service_account_info(info,scopes=['https://www.googleapis.com/auth/drive.readonly'])
    return AuthorizedSession(creds), info.get('client_email','')

def meta(session,file_id):
    r=session.get(f'https://www.googleapis.com/drive/v3/files/{file_id}',params={'fields':'id,name,mimeType,modifiedTime,size,md5Checksum'},timeout=45)
    if r.status_code==404: raise RuntimeError(f'Google Drive file {file_id} is not shared with the service account.')
    r.raise_for_status();return r.json()

def download(session,m):
    fid=m['id'];mime=m.get('mimeType','')
    if mime=='application/vnd.google-apps.spreadsheet':
        r=session.get(f'https://www.googleapis.com/drive/v3/files/{fid}/export',params={'mimeType':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'},timeout=90)
    elif mime=='application/vnd.google-apps.document':
        r=session.get(f'https://www.googleapis.com/drive/v3/files/{fid}/export',params={'mimeType':'text/plain'},timeout=90)
    else:
        r=session.get(f'https://www.googleapis.com/drive/v3/files/{fid}',params={'alt':'media'},timeout=90)
    r.raise_for_status();return r.content

def xlsx_notes(raw,source):
    wb=load_workbook(io.BytesIO(raw),read_only=True,data_only=True)
    notes=[]
    for ws in wb.worksheets:
        for ridx,row in enumerate(ws.iter_rows(values_only=True),start=1):
            cells=[clean(v) for v in row if clean(v)]
            if not cells:continue
            text=' | '.join(cells)
            if len(text)<2:continue
            title=cells[0][:120]
            notes.append({'id':hashlib.sha1(f"{source['id']}:{ws.title}:{ridx}:{text}".encode()).hexdigest()[:18],
                'source_kind':source['kind'],'source_label':source['label'],'file_id':source['id'],'section':ws.title,'location':f'{ws.title}!{ridx}',
                'title':title,'text':text[:5000],'domains':domain_tags(text)})
    return notes

def doc_notes(raw,source):
    text=raw.decode('utf-8','replace').replace('\r\n','\n')
    parts=[p.strip() for p in re.split(r'\n\s*\n+',text) if p.strip()]
    notes=[]
    for i,p in enumerate(parts,start=1):
        p=re.sub(r'\n+','\n',p).strip()
        if len(p)<3:continue
        first=clean(p.split('\n',1)[0])[:120]
        notes.append({'id':hashlib.sha1(f"{source['id']}:{i}:{p}".encode()).hexdigest()[:18],
            'source_kind':source['kind'],'source_label':source['label'],'file_id':source['id'],'section':'Document','location':f'paragraph:{i}',
            'title':first or f'段落 {i}','text':p[:8000],'domains':domain_tags(p)})
    return notes

def sha(raw):return hashlib.sha256(raw).hexdigest()

def previous_hashes():
    if not PUBLIC_OUT.exists():return {}
    try:return json.loads(PUBLIC_OUT.read_text(encoding='utf-8')).get('source_hashes',{})
    except Exception:return {}

def candidate_principles(notes):
    # Deterministic candidates. They remain encrypted; no private text is emitted to public metadata.
    rows=[]
    for n in notes:
        text=n['text']; title=n['title']
        # Short experience/book rows are often already written as a decision rule.
        if n['source_kind'] in ('experience','book') and 6<=len(title)<=110:
            rows.append({'id':'c-'+n['id'],'title':title,'detail':text,'domains':n['domains'],'evidence_ids':[n['id']],
                         'source_kinds':[n['source_kind']],'status':'candidate'})
    return rows[:500]

def llm_synthesize(notes):
    if not OPENAI_API_KEY or not OPENAI_MODEL:return []
    try:
        from openai import OpenAI
        client=OpenAI(api_key=OPENAI_API_KEY)
        groups=defaultdict(list)
        for n in notes:
            for d in n['domains'][:2]:groups[d['name']].append(n)
        all_rules=[]
        schema={'type':'object','properties':{'rules':{'type':'array','maxItems':8,'items':{'type':'object','properties':{
            'title':{'type':'string'},'principle':{'type':'string'},'when':{'type':'string'},
            'steps':{'type':'array','items':{'type':'string'},'maxItems':6},
            'metrics':{'type':'array','items':{'type':'string'},'maxItems':6},
            'traps':{'type':'array','items':{'type':'string'},'maxItems':5},
            'evidence_ids':{'type':'array','items':{'type':'string'},'maxItems':10},
            'tensions':{'type':'array','items':{'type':'string'},'maxItems':4},
            'confidence':{'type':'string','enum':['high','medium','low']}
        },'required':['title','principle','when','steps','metrics','traps','evidence_ids','tensions','confidence'],'additionalProperties':False}}},
        'required':['rules'],'additionalProperties':False}
        for domain,ns in groups.items():
            sample=[];chars=0
            # Mix sources; recent order is not important because these are durable notes.
            for n in ns:
                line=f"[{n['id']}] ({n['source_label']} / {n['section']}) {n['text']}"
                if chars+len(line)>18000:break
                sample.append(line);chars+=len(line)
            if not sample:continue
            prompt=f'''你在整理一个人的长期“工作操作系统”。下面资料来自三类来源：本人工作经验、读书笔记、本人EC/Amazon实战手册。领域：{domain}。
请跨资料合并重复观点，提炼成少而稳的“可执行工作规则”，不要做文章摘要。
每条规则必须回答：什么时候用、核心判断、行动步骤、可观察指标、常见误区。本人经验优先作为实践约束，书本理论作为解释框架；若资料之间冲突，写入 tensions，不要强行统一。
只引用真实存在的 [evidence id]。不要加入资料中完全没有依据的具体数字。
资料：\n'''+"\n".join(sample)
            resp=client.responses.create(model=OPENAI_MODEL,input=prompt,text={'format':{'type':'json_schema','name':'work_system','schema':schema,'strict':True}},store=False)
            data=json.loads(resp.output_text)
            for r in data.get('rules',[]):
                r['domain']=domain;r['id']='r-'+hashlib.sha1((domain+r['title']).encode()).hexdigest()[:14]
                all_rules.append(r)
        return all_rules[:60]
    except Exception as e:
        print('LLM synthesis fallback:',e,file=sys.stderr);return []

def encrypt_payload(payload):
    if not PASSPHRASE or len(PASSPHRASE)<14:
        raise RuntimeError('DASHBOARD_PASSPHRASE must be at least 14 characters for private work-system encryption.')
    salt,iv=secrets.token_bytes(16),secrets.token_bytes(12)
    kdf=PBKDF2HMAC(algorithm=hashes.SHA256(),length=32,salt=salt,iterations=PBKDF2_ITERATIONS)
    key=kdf.derive(PASSPHRASE.encode('utf-8'))
    raw=json.dumps(payload,ensure_ascii=False,separators=(',',':')).encode('utf-8')
    ct=AESGCM(key).encrypt(iv,raw,None)
    b=lambda x:base64.b64encode(x).decode('ascii')
    return {'version':2,'algorithm':'AES-256-GCM','kdf':'PBKDF2-SHA256','iterations':PBKDF2_ITERATIONS,'salt':b(salt),'iv':b(iv),'ciphertext':b(ct)}

def main():
    session,service_email=credentials_session(); source_meta=[]; all_notes=[]; hashes_now={}
    for src in SOURCES:
        m=meta(session,src['id']); raw=download(session,m); digest=sha(raw); hashes_now[src['id']]=digest
        entry={**src,'name':m.get('name') or src['label'],'mimeType':m.get('mimeType'),'modifiedTime':m.get('modifiedTime'),'sha256':digest}
        source_meta.append(entry)
        if m.get('mimeType')=='application/vnd.google-apps.document': all_notes.extend(doc_notes(raw,src))
        else: all_notes.extend(xlsx_notes(raw,src))
        print(f"Fetched {entry['name']}: {len(raw)} bytes")
    if hashes_now==previous_hashes() and ENCRYPTED_OUT.exists():
        print('No Google source changes; keeping current encrypted Work System.');return 0
    rules=llm_synthesize(all_notes); candidates=candidate_principles(all_notes)
    domain_counts=Counter(d['name'] for n in all_notes for d in n['domains'][:1])
    now=datetime.now(timezone.utc).isoformat()
    full={'meta':{'snapshot_at':now,'schema_version':1,'encrypted':True,'service_account':service_email,
                  'synthesis':'openai' if rules else 'deterministic_candidates','source_count':len(source_meta)},
          'sources':source_meta,'notes':all_notes,'candidate_principles':candidates,'rules':rules,
          'domain_counts':[{'name':k,'count':v} for k,v in domain_counts.most_common()]}
    public={'meta':{'snapshot_at':now,'schema_version':1,'encrypted_full_data':True,'source_count':len(source_meta),
                    'note_count':len(all_notes),'rule_count':len(rules),'synthesis':full['meta']['synthesis']},
            'sources':[{'kind':x['kind'],'label':x['label'],'name':x['name'],'modifiedTime':x['modifiedTime']} for x in source_meta],
            'domain_counts':full['domain_counts'],'source_hashes':hashes_now}
    PUBLIC_OUT.parent.mkdir(parents=True,exist_ok=True)
    PUBLIC_OUT.write_text(json.dumps(public,ensure_ascii=False,indent=2),encoding='utf-8')
    ENCRYPTED_OUT.write_text(json.dumps(encrypt_payload(full),ensure_ascii=False),encoding='utf-8')
    print(f'Wrote encrypted Work System: {len(all_notes)} notes, {len(rules)} synthesized rules, {len(candidates)} candidates')
    return 0

if __name__=='__main__': raise SystemExit(main())
